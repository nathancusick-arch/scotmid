"""Scotmid monthly test-purchase report generator.

Run with:
    streamlit run "Scotmid Report Generator.py"

The app accepts the new audit export, the previous LIVE workbook and the
current Scotmid Store DB. It produces an updated internal LIVE workbook, a
client-facing values-only workbook and ready-to-copy email text.
"""

from __future__ import annotations

import hashlib
import io
import re
import zipfile
from collections import Counter, OrderedDict, defaultdict
from copy import copy
from dataclasses import dataclass
from datetime import date, datetime, time
from typing import Any, Iterable
from zoneinfo import ZoneInfo

import pandas as pd
from openpyxl import load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.formula.translate import Translator
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.worksheet.table import TableFormula


GENERATOR_VERSION = "2026.08.12.2"
COMPANY_NAME = "Scotmid Co-operative"


@dataclass
class FieldSpec:
    key: str
    header: str
    source: str | list[str] | None


FIELD_SPECS = [
    FieldSpec("store_code", "Store Number", "site_code"),
    FieldSpec("order", "Order Number", "order_internal_id"),
    FieldSpec("client", "Client Name", "client_name"),
    FieldSpec("visit", "Visit Code", "internal_id"),
    FieldSpec("site_id", "Site Code", "site_internal_id"),
    FieldSpec("deadline", "Order Deadline", "end_date"),
    FieldSpec("responsibility", "Responsibility", "responsibility"),
    FieldSpec("premises", "Premises Name", "site_name"),
    FieldSpec("address1", "Address1", "site_address_1"),
    FieldSpec("address2", "Address2", "site_address_2"),
    FieldSpec("address3", "Address3", "site_address_3"),
    FieldSpec("city", "City", None),
    FieldSpec("postcode", "Post Code", "site_post_code"),
    FieldSpec("start_date", "Start Date", "submitted_date"),
    FieldSpec("end_date", "End Date", "approval_date"),
    FieldSpec("item", "Item to Order", "item_to_order"),
    FieldSpec("visit_date", "Actual Visit Date", "date_of_visit"),
    FieldSpec("visit_time", "Actual Visit Time", "time_of_visit"),
    FieldSpec("ampm", "AMPM", None),
    FieldSpec("result", "Pass-Fail", "primary_result"),
    FieldSpec("result2", "Pass-Fail2", None),
    FieldSpec("abort_reason", "Abort Reason", "site_code"),
    FieldSpec("extra_site_1", "Extra Site 1", "site_code"),
    FieldSpec("extra_site_2", "Extra Site 2", "At which type of till was the purchase made?"),
    FieldSpec("extra_site_3", "Extra Site 3", "At which type of till was the purchase made?"),
    FieldSpec("extra_site_4", "Extra Site 4", "At which type of till was the purchase made?"),
    FieldSpec("extra_site_5", "Extra Site 5", None),
    FieldSpec("extra_site_6", "Extra Site 6", None),
    FieldSpec("extra_site_7", "Extra Site 7", None),
    FieldSpec("extra_site_8", "Extra Site 8", None),
    FieldSpec("extra_site_9", "Extra Site 9", None),
    FieldSpec("extra_site_10", "Extra Site 10", "auditor_gender"),
    FieldSpec("product_type", "What type of alcohol did you purchase?", [
        "What type of alcohol did you purchase?",
        "What type of E-cigarette product did you purchase/attempt to purchase?",
    ]),
    FieldSpec("product_details", "Please give details of the alcohol purchased (brand and size):", [
        "Please give details of the alcohol that you purchased:",
        "Please give details of the cigarettes that you purchased:",
        "Please give details of the e-cig product that you purchased:",
    ]),
    FieldSpec("larger_shop", "Did you make the purchase on its own or as part of a larger shop?",
              "Did you make the purchase on its own or as part of a larger shop?"),
    FieldSpec("age_asked", "At the till / bar / counter, did the person ask you your age during the transaction?",
              "Did the staff member who served you ask your age?"),
    FieldSpec("id_asked", "At the till / bar / counter, did the person (or their supervisor) ask you for ID during the transaction?",
              "Did the staff member who served you ask for ID?"),
    FieldSpec("supervisor_called", "Was a supervisor called at any time during the transaction?",
              "Was a supervisor called at any point during the transaction?"),
    FieldSpec("supervisor_description", "If a supervisor was called, please give an accurate description of the person (hair style and colour / age / build / height / any distinguishing features):",
              "Please accurately describe the supervisor:"),
    FieldSpec("working_alone", "Was the server working entirely alone (i.e. no-one else working in the store)?",
              "Was the staff member who served you working entirely alone?"),
    FieldSpec("eye_contact", "Did the person who served you make eye contact with you during the transaction?",
              "Did the staff member who served you make eye contact with you during the transaction?"),
    FieldSpec("eye_contact_when", "If eye contact was made, when did the person who served you FIRST make eye contact?",
              "When was eye contact first made?"),
    FieldSpec("age_assessed", "Did the server look at you long enough to make an assessment of your age?",
              "Did the staff member who served you look at you long enough to assess your age?"),
    FieldSpec("queue", "How many people were waiting in the queue (if there was no queue, enter 0)?",
              "How many people were in the queue?"),
    FieldSpec("till_type", "", "At which type of till was the purchase made?"),
    FieldSpec("blank_2", "", None),
    FieldSpec("server_gender", "What was the gender of the server?",
              "What was the gender of the staff member who served you?"),
    FieldSpec("server_age", "What was the approximate age of the server?",
              "What was the approximate age of the staff member?"),
    FieldSpec("server_description", "Please describe the hair colour, length and style of the server's hair:",
              "Please accurately describe the staff member who served you:"),
    FieldSpec("name_badge", "Was the server wearing a name badge?",
              "Was the staff member who served you wearing a name badge?"),
    FieldSpec("badge_name", "What is the servers name on the name badge (if visible):",
              "What name was on the name badge?"),
    FieldSpec("receipt_code", "From the receipt, what is the server's name?",
              "Please enter the receipt code shown after the date and time on the receipt:"),
    FieldSpec("operator_code", "Please enter the 'Operator' code:", None),
    FieldSpec("till_code", "Please enter the 'Till' code:", None),
    FieldSpec("receipt_store_code", "Please enter the 'Store' code:", None),
    FieldSpec("store_location", "To help us to identify the site, please describe the surrounding area (i.e. local landmarks, names of stores etc on both sides if possible):",
              "Please describe the location of the store:"),
    FieldSpec("receipt_time", "Please enter the 'Transaction' code:",
              "Please enter the time from the receipt:"),
    FieldSpec("comments", "Did you see or hear anything you think we or our client should know about?", [
        "Please use this space to explain anything unusual about your visit or to clarify any detail of your report:",
    ]),
    FieldSpec("id_confirmation", "", "Please confirm below whether or not you were asked for ID:"),
]


PUBLIC_BASE_SHEETS = [
    "Summary Data", "Historic Data", "Store Performance", "Store Performance (2)",
    "Self-Scan Performance", "Regional Performance", "Regional P.Graphs",
    "Postcode Performance", "Till Type Performance", "Day of Week Performance",
    "Time of Day Performance", "Performance over Time", "Performance over Time Chart",
]
REQUIRED_LIVE_SHEETS = [
    "Checks", "StoreList", "Input", "This Period", "Cumulative", "YTD",
    *PUBLIC_BASE_SHEETS,
]


class ReportGenerationError(ValueError):
    """A friendly validation error suitable for display in Streamlit."""


@dataclass
class Store:
    code: str
    name: str
    region: str
    site_id: str = ""
    postcode: str = ""
    visit_info: str = ""
    status: str = "Active"


@dataclass
class HistoryRecord:
    order: str
    visit: str
    site_id: str
    name: str
    postcode: str
    visit_date: datetime | None
    visit_time: time | None
    result: str
    store_code: str
    till_type: str
    item: str
    raw: list[Any] | None = None


@dataclass
class GenerationResult:
    live_bytes: bytes
    client_bytes: bytes
    zip_bytes: bytes
    live_name: str
    client_name: str
    report_month: date
    product_word: str
    stats: dict[str, Any]
    warnings: list[str]


def _text(value: Any) -> str:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    return str(value).strip()


def _norm(value: Any) -> str:
    return re.sub(r"\s+", " ", _text(value)).strip().casefold()


def _normalise_code(value: Any) -> str:
    text = _text(value)
    if re.fullmatch(r"\d+\.0", text):
        text = text[:-2]
    return text


def _excel_code(value: Any) -> Any:
    code = _normalise_code(value)
    return int(code) if code.isdigit() else code


def _parse_date(value: Any) -> datetime | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.replace(tzinfo=None)
    if isinstance(value, date):
        return datetime(value.year, value.month, value.day)
    parsed = pd.to_datetime(value, dayfirst=True, errors="coerce")
    if pd.isna(parsed):
        return None
    return parsed.to_pydatetime().replace(tzinfo=None)


def _parse_time(value: Any) -> time | None:
    if value in (None, ""):
        return None
    if isinstance(value, time):
        return value.replace(tzinfo=None)
    if isinstance(value, datetime):
        return value.time().replace(tzinfo=None)
    parsed = pd.to_datetime(_text(value), errors="coerce")
    if pd.isna(parsed):
        return None
    return parsed.to_pydatetime().time().replace(tzinfo=None)


def _month_start(value: date | datetime) -> date:
    return date(value.year, value.month, 1)


def _add_months(value: date, months: int) -> date:
    index = value.year * 12 + value.month - 1 + months
    return date(index // 12, index % 12 + 1, 1)


def _natural_key(value: str) -> tuple:
    return tuple(int(part) if part.isdigit() else part.casefold()
                 for part in re.split(r"(\d+)", value))


def _clean_store_name(value: Any, code: str) -> str:
    name = _text(value).replace("_x000D_", " ").replace("\n", " ")
    name = re.sub(r"\s+", " ", name).strip(" ,-_")
    name = re.sub(r"^Scotmid\s*,\s*", "", name, flags=re.I)
    return name or f"Store {code}"


def _result(value: Any) -> str:
    result = _norm(value)
    return result if result in {"pass", "fail"} else ""


def _display_result(value: Any) -> str:
    result = _result(value)
    return "P" if result == "pass" else ("fail" if result == "fail" else "-")


def _product_family(value: Any) -> str | None:
    item = _norm(value).replace("–", "-")
    if item.startswith("alcohol"):
        return "alcohol"
    if item.startswith("cigarette") or item.startswith("tobacco"):
        return "cigarettes"
    if item.startswith("e-cig") or item.startswith("ecig") or item.startswith("vape"):
        return "e-cig"
    return None


def _value_from_row(row: pd.Series, source: str | list[str] | None,
                    lookup: dict[str, str]) -> str:
    if source is None:
        return ""
    sources = source if isinstance(source, list) else [source]
    values: list[str] = []
    for candidate in sources:
        actual = candidate if candidate in row.index else lookup.get(_norm(candidate))
        if actual is not None:
            value = _text(row.get(actual))
            if value and value not in values:
                values.append(value)
    return " | ".join(values)


def _typed_value(key: str, value: Any) -> Any:
    if key in {"deadline", "start_date", "end_date", "visit_date"}:
        return _parse_date(value)
    if key in {"visit_time", "receipt_time"}:
        return _parse_time(value)
    if key in {"store_code", "abort_reason", "extra_site_1"}:
        return _excel_code(value) if _normalise_code(value) else None
    if key == "queue" and re.fullmatch(r"\d+", _text(value)):
        return int(_text(value))
    return _text(value)


def _increment_order_id(value: Any) -> str:
    text = _text(value)
    match = re.match(r"^(.*?)(\d+)$", text)
    return f"{match.group(1)}{int(match.group(2)) + 1}" if match else text


def map_audit_export(csv_bytes: bytes) -> tuple[list[dict[str, Any]], date, str, dict[str, int], list[str]]:
    try:
        frame = pd.read_csv(io.BytesIO(csv_bytes), dtype=str, encoding="utf-8-sig").fillna("")
    except Exception as exc:
        raise ReportGenerationError(f"The audit export could not be read as CSV: {exc}") from exc

    required = {"item_to_order", "primary_result", "date_of_visit", "internal_id", "site_code"}
    missing = sorted(required - set(frame.columns))
    if missing:
        raise ReportGenerationError("The audit export is missing required column(s): " + ", ".join(missing))

    abort_mask = frame["primary_result"].map(_norm).eq("abort")
    product_families = frame["item_to_order"].map(_product_family)
    relevant_mask = product_families.notna()
    filtered = frame.loc[~abort_mask & relevant_mask].copy()
    if filtered.empty:
        raise ReportGenerationError("No Alcohol, Cigarettes or E-Cig visits remain after aborts are removed.")

    dates = filtered["date_of_visit"].map(_parse_date)
    if dates.isna().any():
        bad_index = dates[dates.isna()].index[0]
        raise ReportGenerationError(
            f"An included audit has an invalid Actual Visit Date (visit {_text(frame.loc[bad_index, 'internal_id'])})."
        )
    month_counts = Counter(_month_start(value) for value in dates)
    report_month, report_count = month_counts.most_common(1)[0]
    product_counts = Counter(_product_family(value) for value in filtered["item_to_order"])
    product_word, _ = product_counts.most_common(1)[0]
    lookup = {_norm(column): column for column in frame.columns}

    records: list[dict[str, Any]] = []
    for _, source_row in filtered.iterrows():
        record: dict[str, Any] = {}
        for spec in FIELD_SPECS:
            record[spec.key] = _typed_value(
                spec.key, _value_from_row(source_row, spec.source, lookup)
            )
        # The legacy Scotmid workbook uses the following order number for odd
        # calendar months; this preserves the behaviour of the former mapper.
        if record["visit_date"] and record["visit_date"].month % 2 == 1:
            record["order"] = _increment_order_id(record["order"])
        record["values"] = [record[spec.key] for spec in FIELD_SPECS]
        records.append(record)

    visits = [_text(record["visit"]) for record in records]
    duplicates = [visit for visit, count in Counter(visits).items() if visit and count > 1]
    if duplicates:
        raise ReportGenerationError(f"The audit export contains duplicate visit code {duplicates[0]}.")

    warnings: list[str] = []
    if len(month_counts) > 1:
        others = ", ".join(f"{month:%B %Y} ({count})" for month, count in sorted(month_counts.items())
                           if month != report_month)
        warnings.append(
            f"The export spans multiple visit months. {report_month:%B %Y} was selected from "
            f"{report_count} row(s); other month(s): {others}."
        )
    if len(product_counts) > 1:
        warnings.append(
            f"The export contains more than one product family. The email uses the most common: {product_word}."
        )
    stats = {
        "export_rows": len(frame),
        "aborts_removed": int(abort_mask.sum()),
        "unrelated_rows_removed": int((~relevant_mask & ~abort_mask).sum()),
        "included_rows": len(records),
    }
    return records, report_month, product_word, stats, warnings


def _load_live_workbook(workbook_bytes: bytes):
    try:
        workbook = load_workbook(io.BytesIO(workbook_bytes), data_only=False, keep_links=False)
    except Exception as exc:
        raise ReportGenerationError(f"The previous LIVE workbook could not be read: {exc}") from exc
    missing = [name for name in REQUIRED_LIVE_SHEETS if name not in workbook.sheetnames]
    if missing:
        workbook.close()
        raise ReportGenerationError(
            "The previous workbook is not a compatible Scotmid LIVE report. Missing tab(s): "
            + ", ".join(missing)
        )
    return workbook


def read_store_database(store_bytes: bytes) -> OrderedDict[str, Store]:
    try:
        workbook = load_workbook(io.BytesIO(store_bytes), data_only=True, read_only=True, keep_links=False)
    except Exception as exc:
        raise ReportGenerationError(f"The Store DB could not be read: {exc}") from exc
    sheet = workbook["MASTER IMPORT -  NEW SYSTEM"] if "MASTER IMPORT -  NEW SYSTEM" in workbook.sheetnames else workbook.active
    header_values = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = {_norm(value): index for index, value in enumerate(header_values)}
    aliases = {
        "code": ["code", "store code", "store number"],
        "name": ["name", "store name"],
        "region": ["area", "region"],
        "site_id": ["site internal id", "site id"],
        "postcode": ["post code", "postcode"],
        "visit_info": ["visit info"],
    }

    def index_for(field: str, required: bool = False) -> int | None:
        for alias in aliases[field]:
            if alias in headers:
                return headers[alias]
        if required:
            raise ReportGenerationError(f"The Store DB is missing the {field.replace('_', ' ')} column.")
        return None

    indices = {field: index_for(field, field in {"code", "name", "region"}) for field in aliases}
    stores: OrderedDict[str, Store] = OrderedDict()
    for values in sheet.iter_rows(min_row=2, values_only=True):
        code = _normalise_code(values[indices["code"]])
        if not code:
            continue
        if code in stores:
            workbook.close()
            raise ReportGenerationError(f"The Store DB contains duplicate store code {code}.")
        region = _normalise_code(values[indices["region"]]) or "Unmapped"
        site_id = _text(values[indices["site_id"]]) if indices["site_id"] is not None else ""
        postcode = _text(values[indices["postcode"]]) if indices["postcode"] is not None else ""
        visit_info = _text(values[indices["visit_info"]]) if indices["visit_info"] is not None else ""
        stores[code] = Store(
            code=code,
            name=_clean_store_name(values[indices["name"]], code),
            region=region,
            site_id=site_id,
            postcode=postcode,
            visit_info=visit_info,
        )
    workbook.close()
    if not stores:
        raise ReportGenerationError("No stores were found in the Store DB.")
    return stores


def extract_previous_stores(workbook) -> OrderedDict[str, Store]:
    sheet = workbook["StoreList"]
    stores: OrderedDict[str, Store] = OrderedDict()
    for row in range(2, sheet.max_row + 1):
        code = _normalise_code(sheet.cell(row, 1).value)
        if not code:
            continue
        region = _normalise_code(sheet.cell(row, 3).value) or "closed"
        stores[code] = Store(
            code=code,
            name=_clean_store_name(sheet.cell(row, 2).value, code),
            region=region,
            status="Closed" if region.casefold() == "closed" else "Previous",
        )
    return stores


def build_hierarchy(database: OrderedDict[str, Store], previous: OrderedDict[str, Store],
                    records: list[dict[str, Any]]) -> tuple[OrderedDict[str, Store], list[str]]:
    hierarchy: OrderedDict[str, Store] = OrderedDict()
    for code in previous:
        if code in database:
            hierarchy[code] = database[code]
    for code, store in database.items():
        hierarchy.setdefault(code, store)
    for code, old in previous.items():
        if code not in database:
            hierarchy[code] = Store(code, old.name, "closed", status="Closed")

    warnings: list[str] = []
    for record in records:
        code = _normalise_code(record["store_code"])
        if code and code not in hierarchy:
            name = _clean_store_name(record["premises"], code)
            hierarchy[code] = Store(code, name, "Unmapped", _text(record["site_id"]), status="Unmapped")
            warnings.append(f"Store {code} ({name}) is present in the audit export but not in the Store DB.")
    return hierarchy, warnings


def _last_data_row(sheet, column: int, first_row: int) -> int:
    rows = [row for (row, col), cell in sheet._cells.items()
            if col == column and row >= first_row and cell.value not in (None, "")]
    return max(rows, default=first_row - 1)


def _copy_row_style(sheet, source_row: int, target_row: int, max_column: int) -> None:
    sheet.row_dimensions[target_row].height = sheet.row_dimensions[source_row].height
    for column in range(1, max_column + 1):
        source = sheet.cell(source_row, column)
        target = sheet.cell(target_row, column)
        if source.has_style:
            target._style = copy(source._style)
        target.number_format = source.number_format
        target.alignment = copy(source.alignment)
        target.protection = copy(source.protection)


def _snapshot_row_style(sheet, row: int, max_column: int) -> tuple[Any, bool, list[Any]]:
    """Capture a row's exact layout before its former position is reused."""
    return (
        sheet.row_dimensions[row].height,
        sheet.row_dimensions[row].hidden,
        [copy(sheet.cell(row, column)._style) for column in range(1, max_column + 1)],
    )


def _apply_row_style(sheet, row: int, snapshot: tuple[Any, bool, list[Any]]) -> None:
    height, hidden, styles = snapshot
    sheet.row_dimensions[row].height = height
    sheet.row_dimensions[row].hidden = hidden
    for column, style in enumerate(styles, 1):
        sheet.cell(row, column)._style = copy(style)


def _find_row(sheet, value: str, column: int = 1, start: int = 1) -> int | None:
    wanted = _norm(value)
    for row in range(start, sheet.max_row + 1):
        if _norm(sheet.cell(row, column).value) == wanted:
            return row
    return None


def _trim_rows(sheet, last_row: int) -> None:
    """Remove unused formatted rows so they do not remain in client tables."""
    if sheet.max_row > last_row:
        sheet.delete_rows(last_row + 1, sheet.max_row - last_row)


def _capture_formulas(sheet, row: int, start_column: int, end_column: int) -> dict[int, str]:
    return {
        column: sheet.cell(row, column).value
        for column in range(start_column, end_column + 1)
        if isinstance(sheet.cell(row, column).value, str)
        and sheet.cell(row, column).value.startswith("=")
    }


def _translate_formula(formula: str, source_row: int, target_row: int, column: int) -> str:
    coordinate = f"{get_column_letter(column)}{source_row}"
    target = f"{get_column_letter(column)}{target_row}"
    return Translator(formula, origin=coordinate).translate_formula(target)


def update_this_period(sheet, records: list[dict[str, Any]]) -> int:
    old_last = _last_data_row(sheet, 1, 4)
    new_last = 3 + len(records)
    source_row = 4
    formulas = _capture_formulas(sheet, source_row, 239, 252)
    clear_to = max(old_last, new_last)
    for row in range(4, clear_to + 1):
        if row > sheet.max_row:
            _copy_row_style(sheet, source_row, row, 252)
        for column in range(1, 253):
            sheet.cell(row, column).value = None
    for offset, record in enumerate(records):
        row = 4 + offset
        for column, value in enumerate(record["values"], 1):
            sheet.cell(row, column).value = value
        for column, formula in formulas.items():
            sheet.cell(row, column).value = _translate_formula(formula, source_row, row, column)
    end = max(4, new_last)
    sheet["B2"] = f"=COUNTA(B4:B{end})"
    sheet["IR2"] = f"=COUNTA(IR4:IR{end})"
    sheet.auto_filter.ref = f"A3:IR{end}"
    return new_last


def update_cumulative(sheet, records: list[dict[str, Any]]) -> int:
    last_row = _last_data_row(sheet, 1, 4)
    formula_row = last_row
    formulas = _capture_formulas(sheet, formula_row, 236, 256)
    visit_rows = {
        _text(sheet.cell(row, 3).value): row
        for row in range(4, last_row + 1)
        if _text(sheet.cell(row, 3).value)
    }
    for record in records:
        visit = _text(record["visit"])
        target_row = visit_rows.get(visit)
        if target_row is None:
            last_row += 1
            target_row = last_row
            _copy_row_style(sheet, formula_row, target_row, 257)
        for column in range(1, 59):
            sheet.cell(target_row, column).value = None
        for column, value in enumerate(record["values"][1:], 1):
            sheet.cell(target_row, column).value = value
        for column, formula in formulas.items():
            sheet.cell(target_row, column).value = _translate_formula(
                formula, formula_row, target_row, column
            )
    end = max(4, last_row)
    sheet["A2"] = f"=COUNTA(A4:A{end})"
    sheet["IU2"] = f"=COUNTA(IU4:IU{end})"
    sheet["IV2"] = f"=COUNTA(IV4:IV{end})"
    sheet.auto_filter.ref = f"A3:IW{end}"
    return last_row


def extract_history(sheet, report_year: int) -> tuple[list[HistoryRecord], list[HistoryRecord]]:
    last_row = _last_data_row(sheet, 1, 4)
    history: list[HistoryRecord] = []
    ytd: list[HistoryRecord] = []
    for row in range(4, last_row + 1):
        visit_date = _parse_date(sheet.cell(row, 16).value)
        record = HistoryRecord(
            order=_text(sheet.cell(row, 1).value),
            visit=_text(sheet.cell(row, 3).value),
            site_id=_text(sheet.cell(row, 4).value),
            name=_text(sheet.cell(row, 7).value),
            postcode=_text(sheet.cell(row, 11).value),
            visit_date=visit_date,
            visit_time=_parse_time(sheet.cell(row, 17).value),
            result=_result(sheet.cell(row, 19).value),
            store_code=_normalise_code(sheet.cell(row, 22).value),
            till_type=_text(sheet.cell(row, 23).value),
            item=_text(sheet.cell(row, 15).value),
        )
        history.append(record)
        if visit_date and visit_date.year == report_year:
            record.raw = [sheet.cell(row, column).value for column in range(1, 59)]
            ytd.append(record)
    return history, ytd


def update_ytd(sheet, records: list[HistoryRecord]) -> int:
    old_last = _last_data_row(sheet, 1, 4)
    new_last = 3 + len(records)
    source_row = min(max(old_last, 4), sheet.max_row)
    formulas = _capture_formulas(sheet, source_row, 236, 256)
    for row in range(4, max(old_last, new_last) + 1):
        if row > sheet.max_row:
            _copy_row_style(sheet, source_row, row, 256)
        for column in range(1, 257):
            sheet.cell(row, column).value = None
    for offset, record in enumerate(records):
        row = 4 + offset
        for column, value in enumerate(record.raw or [], 1):
            sheet.cell(row, column).value = value
        for column, formula in formulas.items():
            sheet.cell(row, column).value = _translate_formula(formula, source_row, row, column)
    end = max(4, new_last)
    sheet["A2"] = f"=COUNTA(A4:A{end})"
    sheet["IU2"] = f"=COUNTA(IU4:IU{end})"
    sheet["IV2"] = f"=COUNTA(IV4:IV{end})"
    sheet.auto_filter.ref = f"A3:IV{end}"
    return new_last


def update_store_list(sheet, hierarchy: OrderedDict[str, Store]) -> int:
    old_last = _last_data_row(sheet, 1, 2)
    new_last = 1 + len(hierarchy)
    source_row = 2
    for row in range(2, max(old_last, new_last) + 1):
        if row > sheet.max_row:
            _copy_row_style(sheet, source_row, row, 3)
        for column in range(1, 4):
            sheet.cell(row, column).value = None
    for row, store in enumerate(hierarchy.values(), 2):
        sheet.cell(row, 1).value = _excel_code(store.code)
        sheet.cell(row, 2).value = store.name
        sheet.cell(row, 3).value = _excel_code(store.region) if store.region.isdigit() else store.region
    sheet.auto_filter.ref = f"A1:C{new_last}"
    return new_last


def _records_in_month(records: Iterable[HistoryRecord], month: date) -> list[HistoryRecord]:
    next_month = _add_months(month, 1)
    return [record for record in records if record.visit_date
            and month <= record.visit_date.date() < next_month]


def update_input(sheet, history: list[HistoryRecord], current: list[dict[str, Any]],
                 report_month: date) -> list[date]:
    months = [_add_months(report_month, offset) for offset in range(-11, 1)]
    current_orders = Counter(_text(record["order"]) for record in current)
    for index, month in enumerate(months, 18):
        month_records = _records_in_month(history, month)
        order_counts = Counter(record.order for record in month_records if record.order)
        order = order_counts.most_common(1)[0][0] if order_counts else ""
        sheet.cell(index, 1).value = order
        sheet.cell(index, 2).value = datetime(month.year, month.month, 10)
        sheet.cell(index, 4).value = len(month_records)
        sheet.cell(index, 5).value = current_orders.get(order, 0)
    sheet["D16"] = sum(sheet.cell(row, 4).value or 0 for row in range(18, 30))
    sheet["E16"] = len(current)
    return months


def _metric(records: Iterable[Any]) -> list[Any]:
    records = list(records)
    results = [_result(record.result if isinstance(record, HistoryRecord) else record["result"])
               for record in records]
    completed = sum(result in {"pass", "fail"} for result in results)
    fails = sum(result == "fail" for result in results)
    passes = sum(result == "pass" for result in results)
    return [len(records), completed, fails, passes, passes / completed if completed else "-"]


def _copy_cell_style(source, target) -> None:
    if source.has_style:
        target._style = copy(source._style)
    target.number_format = source.number_format
    target.alignment = copy(source.alignment)
    target.protection = copy(source.protection)


def _ensure_rows(sheet, needed_row: int, style_row: int, max_column: int) -> None:
    for row in range(sheet.max_row + 1, needed_row + 1):
        _copy_row_style(sheet, style_row, row, max_column)


def _clear_values(sheet, min_row: int, max_row: int, min_col: int, max_col: int) -> None:
    for row in range(min_row, max_row + 1):
        for column in range(min_col, max_col + 1):
            cell = sheet.cell(row, column)
            if cell.__class__.__name__ != "MergedCell":
                cell.value = None


def _strip_formulas(sheet) -> None:
    if not hasattr(sheet, "iter_rows"):
        return
    for row in sheet.iter_rows():
        for cell in row:
            if cell.data_type == "f":
                cell.value = None


def _set_report_date(sheet, report_month: date) -> None:
    if sheet.max_column >= 11:
        sheet.cell(2, sheet.max_column).value = COMPANY_NAME
        sheet.cell(3, sheet.max_column).value = report_month.strftime("01/%m/%Y")


def populate_summary(sheet, current: list[dict[str, Any]]) -> None:
    _strip_formulas(sheet)
    old_last = _last_data_row(sheet, 1, 8)
    new_last = 7 + len(current)
    _ensure_rows(sheet, new_last, 8, 24)
    _clear_values(sheet, 8, max(old_last, new_last), 1, 24)
    data_style = _snapshot_row_style(sheet, 8, 24)
    completed = [record for record in current if _result(record["result"])]
    sheet["B3"] = len({_normalise_code(record["store_code"]) for record in current if _normalise_code(record["store_code"])})
    sheet["B4"] = len(completed)
    passes = sum(_result(record["result"]) == "pass" for record in completed)
    sheet["B5"] = passes / len(completed) if completed else "-"
    for row, record in enumerate(current, 8):
        _apply_row_style(sheet, row, data_style)
        visit_date = record["visit_date"]
        values = [
            record["store_code"], record["premises"], record["postcode"], visit_date,
            visit_date.strftime("%a").replace("Tue", "Tues").replace("Thu", "Thur") if visit_date else "",
            record["visit_time"], _result(record["result"]).upper(), record["product_type"],
            record["product_details"], record["larger_shop"], record["age_asked"], record["id_asked"],
            record["supervisor_called"], record["working_alone"], record["eye_contact"],
            record["eye_contact_when"], record["age_assessed"], record["queue"], record["name_badge"],
            record["receipt_code"], record["receipt_time"], record["id_confirmation"],
            record["store_location"], record["comments"],
        ]
        for column, value in enumerate(values, 1):
            sheet.cell(row, column).value = value
    _trim_rows(sheet, max(7, new_last))


def populate_historic(sheet, history: list[HistoryRecord]) -> None:
    _strip_formulas(sheet)
    old_last = _last_data_row(sheet, 1, 4)
    new_last = 3 + len(history)
    _ensure_rows(sheet, new_last, 4, 5)
    _clear_values(sheet, 4, max(old_last, new_last), 1, 5)
    for row, record in enumerate(history, 4):
        values = [_excel_code(record.store_code), record.name, record.postcode,
                  record.visit_date, _result(record.result).upper()]
        for column, value in enumerate(values, 1):
            sheet.cell(row, column).value = value
    sheet.auto_filter.ref = f"A3:E{max(4, new_last)}"


def populate_store_performance(sheet, league_sheet, hierarchy: OrderedDict[str, Store],
                               current: list[dict[str, Any]], ytd: list[HistoryRecord],
                               report_month: date) -> int:
    _strip_formulas(sheet)
    _strip_formulas(league_sheet)
    stores = list(hierarchy.values())
    old_total_row = _find_row(sheet, "Total", 1, 7) or sheet.max_row
    store_style = _snapshot_row_style(sheet, 7, 12)
    gap_style = _snapshot_row_style(sheet, max(7, old_total_row - 1), 12)
    total_style = _snapshot_row_style(sheet, old_total_row, 12)
    last_store_row = 6 + len(stores)
    gap_row = last_store_row + 1
    total_row = last_store_row + 2
    _ensure_rows(sheet, total_row, 7, 12)
    _clear_values(sheet, 7, max(sheet.max_row, total_row), 1, 12)
    current_by = defaultdict(list)
    ytd_by = defaultdict(list)
    for record in current:
        current_by[_normalise_code(record["store_code"])].append(record)
    for record in ytd:
        ytd_by[record.store_code].append(record)
    for row, store in enumerate(stores, 7):
        _apply_row_style(sheet, row, store_style)
        values = [_excel_code(store.code), store.name, *_metric(current_by[store.code]), *_metric(ytd_by[store.code])]
        for column, value in enumerate(values, 1):
            sheet.cell(row, column).value = value
    _apply_row_style(sheet, gap_row, gap_style)
    _apply_row_style(sheet, total_row, total_style)
    total_values = ["Total", "", *_metric(current), *_metric(ytd)]
    for column, value in enumerate(total_values, 1):
        sheet.cell(total_row, column).value = value
        font = copy(sheet.cell(total_row, column).font)
        font.bold = True
        sheet.cell(total_row, column).font = font
    sheet["L2"] = COMPANY_NAME
    sheet["L3"] = report_month.strftime("01/%m/%Y")
    _trim_rows(sheet, total_row)

    league_last = 5 + len(stores)
    league_style = _snapshot_row_style(league_sheet, 6, 3)
    _ensure_rows(league_sheet, league_last, 6, 3)
    _clear_values(league_sheet, 6, max(league_sheet.max_row, league_last), 1, 3)
    for row, store in enumerate(stores, 6):
        _apply_row_style(league_sheet, row, league_style)
        league_sheet.cell(row, 1).value = _excel_code(store.code)
        league_sheet.cell(row, 2).value = store.name
        league_sheet.cell(row, 3).value = _metric(ytd_by[store.code])[2]
    league_sheet.auto_filter.ref = f"A5:C{max(6, league_last)}"
    _trim_rows(league_sheet, max(5, league_last))
    return total_row


def _normalise_till(value: Any) -> str:
    text = _norm(value)
    if "self" in text and "scan" in text:
        return "Self-Scan Till"
    if "kiosk" in text:
        return "Kiosk"
    return "Normal Till" if text else ""


def populate_self_scan(sheet, hierarchy: OrderedDict[str, Store], history: list[HistoryRecord],
                       months: list[date]) -> None:
    _strip_formulas(sheet)
    table_by_name = {table.name: table for table in sheet.tables.values()}
    right_table = table_by_name.get("Table1")
    totals_table = table_by_name.get("Table2")
    if right_table is None or totals_table is None:
        raise ReportGenerationError("The Self-Scan Performance template tables are missing.")
    _, _, _, old_right_total = range_boundaries(right_table.ref)
    _, old_totals_start, _, old_totals_end = range_boundaries(totals_table.ref)
    data_style = _snapshot_row_style(sheet, 6, 20)
    right_total_style = _snapshot_row_style(sheet, old_right_total, 20)
    blank_style = _snapshot_row_style(sheet, min(old_right_total + 1, old_totals_start - 1), 20)
    bottom_styles = [
        _snapshot_row_style(sheet, row, 20)
        for row in range(old_totals_start, old_totals_end + 1)
    ]

    rolling = [record for record in history if record.visit_date and months[0] <= record.visit_date.date() < _add_months(months[-1], 1)]
    codes = [
        code for code, store in hierarchy.items()
        if store.status == "Active" and "self scan till" in _norm(store.visit_info)
    ]
    code_set = set(codes)
    first_row = 6
    last_store_row = first_row + len(codes) - 1
    right_total_row = last_store_row + 1
    total_start = last_store_row + 4
    total_end = total_start + 2
    _ensure_rows(sheet, total_end, 6, 20)
    _clear_values(sheet, 4, max(sheet.max_row, total_end), 1, 20)
    for offset, month in enumerate(months, 3):
        month_records = _records_in_month(rolling, month)
        orders = Counter(record.order for record in month_records if record.order)
        sheet.cell(4, offset).value = orders.most_common(1)[0][0] if orders else ""
        sheet.cell(5, offset).value = datetime(month.year, month.month, 10)
    sheet["A5"] = "Store No"
    sheet["B5"] = "Store Name"
    sheet["R5"] = "Total Tests"
    sheet["S5"] = "Total Passes"
    sheet["T5"] = "%"
    for row, code in enumerate(codes, first_row):
        _apply_row_style(sheet, row, data_style)
        store = hierarchy[code]
        sheet.cell(row, 1).value = _excel_code(code)
        sheet.cell(row, 2).value = store.name
        tests = passes = 0
        for offset, month in enumerate(months, 3):
            records = [record for record in _records_in_month(rolling, month)
                       if record.store_code == code and _normalise_till(record.till_type) == "Self-Scan Till"]
            value = "-"
            if records:
                tests += sum(bool(_result(record.result)) for record in records)
                passes += sum(_result(record.result) == "pass" for record in records)
                value = "fail" if any(_result(record.result) == "fail" for record in records) else "P"
            sheet.cell(row, offset).value = value
        sheet.cell(row, 18).value = tests if tests else "-"
        sheet.cell(row, 19).value = passes if tests else "-"
        sheet.cell(row, 20).value = passes / tests if tests else "-"

    _apply_row_style(sheet, right_total_row, right_total_style)
    total_tests = sum(
        bool(_result(record.result)) for record in rolling
        if record.store_code in code_set and _normalise_till(record.till_type) == "Self-Scan Till"
    )
    total_passes = sum(
        _result(record.result) == "pass" for record in rolling
        if record.store_code in code_set and _normalise_till(record.till_type) == "Self-Scan Till"
    )
    sheet.cell(right_total_row, 18).value = total_tests if total_tests else "-"
    sheet.cell(right_total_row, 19).value = total_passes if total_tests else "-"
    sheet.cell(right_total_row, 20).value = total_passes / total_tests if total_tests else "-"

    for row in range(right_total_row + 1, total_start):
        _apply_row_style(sheet, row, blank_style)
    for index, label in enumerate(["Total Tests", "Total Passes", "%"]):
        row = total_start + index
        _apply_row_style(sheet, row, bottom_styles[index])
        sheet.cell(row, 1).value = label
        sheet.cell(row, 2).value = "="
        sheet.cell(row, 2).data_type = "s"
        for offset, month in enumerate(months, 3):
            records = [record for record in _records_in_month(rolling, month)
                       if record.store_code in code_set
                       and _normalise_till(record.till_type) == "Self-Scan Till"
                       and _result(record.result)]
            tests = len(records)
            passes = sum(_result(record.result) == "pass" for record in records)
            sheet.cell(row, offset).value = [tests, passes, passes / tests if tests else "-"][index]
            if not tests:
                sheet.cell(row, offset).value = "-"

    right_table.ref = f"R5:T{right_total_row}"
    right_table.totalsRowCount = 1
    right_table.totalsRowShown = True
    right_formulas = [
        (
            'COUNTIF(C6:N6,"P")+COUNTIF(C6:N6,"Fail")',
            f"SUM(R6:R{last_store_row})",
        ),
        ('COUNTIF(C6:N6,"P")', f"SUM(S6:S{last_store_row})"),
        (
            'IF(Table1[[#This Row],[Total Tests]]=0,"-",S6/R6)',
            f'IF(R{right_total_row}=0,"-",S{right_total_row}/R{right_total_row})',
        ),
    ]
    for column, (calculated, total) in zip(right_table.tableColumns, right_formulas):
        column.calculatedColumnFormula = TableFormula(attr_text=calculated)
        column.totalsRowFormula = TableFormula(attr_text=total)
        column.totalsRowFunction = "custom"
    totals_table.ref = f"A{total_start}:N{total_end}"

    sheet.conditional_formatting._cf_rules.clear()
    sheet.conditional_formatting.add(
        f"C6:N{last_store_row}",
        FormulaRule(formula=['LOWER(C6)="fail"'], font=Font(color="FFFF0000")),
    )
    _trim_rows(sheet, total_end)


def populate_metric_sheet(sheet, labels: list[Any], current: list[dict[str, Any]],
                          ytd: list[HistoryRecord], current_key, ytd_key,
                          report_month: date) -> None:
    _strip_formulas(sheet)
    old_total = _last_data_row(sheet, 1, 7)
    first_row = 7
    total_row = first_row + len(labels) + 1
    _ensure_rows(sheet, total_row, first_row, 11)
    _clear_values(sheet, first_row, max(old_total, total_row), 1, 11)
    current_groups = defaultdict(list)
    ytd_groups = defaultdict(list)
    for record in current:
        current_groups[current_key(record)].append(record)
    for record in ytd:
        ytd_groups[ytd_key(record)].append(record)
    for row, label in enumerate(labels, first_row):
        values = [label, *_metric(current_groups[label]), *_metric(ytd_groups[label])]
        for column, value in enumerate(values, 1):
            sheet.cell(row, column).value = value
    totals = ["Total", *_metric(current), *_metric(ytd)]
    for column, value in enumerate(totals, 1):
        sheet.cell(total_row, column).value = value
    _set_report_date(sheet, report_month)


def _postcode_region(value: Any) -> str:
    match = re.match(r"([A-Z]{1,2})", _text(value).upper())
    return match.group(1) if match else "Other"


def _day_label(value: Any) -> str:
    parsed = _parse_date(value)
    if not parsed:
        return ""
    return ["Mon", "Tues", "Wed", "Thur", "Fri", "Sat", "Sun"][parsed.weekday()]


def _hour(value: Any) -> int | None:
    parsed = _parse_time(value)
    return parsed.hour if parsed else None


def populate_performance_over_time(sheet, history: list[HistoryRecord], months: list[date]) -> None:
    _strip_formulas(sheet)
    sheet["M2"] = COMPANY_NAME
    sheet["A5"] = "Name"
    sheet["A6"] = "Overall Pass Rate"
    for column, month in enumerate(months, 2):
        records = [record for record in _records_in_month(history, month) if _result(record.result)]
        sheet.cell(5, column).value = month.strftime("%b %Y")
        sheet.cell(6, column).value = _metric(records)[4]


def _region_month_metrics(history: list[HistoryRecord], hierarchy: OrderedDict[str, Store],
                          region: str, month: date) -> tuple[int, int, Any]:
    records = [record for record in _records_in_month(history, month)
               if record.store_code in hierarchy and hierarchy[record.store_code].region == region
               and _result(record.result)]
    passes = sum(_result(record.result) == "pass" for record in records)
    fails = sum(_result(record.result) == "fail" for record in records)
    return passes, fails, passes / (passes + fails) * 100 if passes + fails else ""


def populate_regional_performance(sheet, graph_sheet, hierarchy: OrderedDict[str, Store],
                                  history: list[HistoryRecord], months: list[date],
                                  regions: list[str]) -> None:
    _strip_formulas(sheet)
    blank_style = _snapshot_row_style(sheet, 35, 36)
    for block, block_months in [(0, months[:6]), (1, months[6:])]:
        header_row = 4 + block * 16
        date_row = 5 + block * 16
        label_row = 6 + block * 16
        data_start = 7 + block * 16
        for offset, month in enumerate(block_months):
            column = 5 + offset * 3
            month_records = _records_in_month(history, month)
            orders = Counter(record.order for record in month_records if record.order)
            sheet.cell(header_row, column).value = orders.most_common(1)[0][0] if orders else ""
            sheet.cell(date_row, column).value = datetime(month.year, month.month, 10)
            for delta, value in enumerate(["Pass", "Fail", "%"]):
                sheet.cell(label_row, column + delta).value = value
        sheet.cell(date_row, 4).value = "Region"
        _clear_values(sheet, data_start, data_start + 11, 4, 22)
        for row, region in enumerate(regions, data_start):
            sheet.cell(row, 4).value = _excel_code(region) if region.isdigit() else region
            for offset, month in enumerate(block_months):
                column = 5 + offset * 3
                for delta, value in enumerate(_region_month_metrics(history, hierarchy, region, month)):
                    sheet.cell(row, column + delta).value = value
        total_row = data_start + len(regions)
        sheet.cell(total_row, 4).value = "Total"
        for offset, month in enumerate(block_months):
            column = 5 + offset * 3
            month_records = [record for record in _records_in_month(history, month) if _result(record.result)]
            passes = sum(_result(record.result) == "pass" for record in month_records)
            fails = sum(_result(record.result) == "fail" for record in month_records)
            sheet.cell(total_row, column).value = passes
            sheet.cell(total_row, column + 1).value = fails
            sheet.cell(total_row, column + 2).value = passes / (passes + fails) * 100 if passes + fails else ""
        for row in range(total_row + 1, data_start + 12):
            _apply_row_style(sheet, row, blank_style)
            _clear_values(sheet, row, row, 1, 36)

    sheet["F38"] = "Rolling 12 months"
    sheet["C39"] = "Region"
    sheet["C40"] = "Pass rate"
    _clear_values(sheet, 39, 40, 4, 15)
    rolling = [record for record in history if record.visit_date
               and months[0] <= record.visit_date.date() < _add_months(months[-1], 1)]
    for column, region in enumerate(regions, 4):
        records = [record for record in rolling if record.store_code in hierarchy
                   and hierarchy[record.store_code].region == region and _result(record.result)]
        sheet.cell(39, column).value = _excel_code(region) if region.isdigit() else region
        rate = _metric(records)[4]
        sheet.cell(40, column).value = rate if rate != "-" else ""

    if graph_sheet._charts:
        chart = graph_sheet._charts[0]
        category_source = chart.series[0].cat
        if getattr(category_source, "strRef", None) is not None:
            category_source.strRef.f = f"'Regional Performance'!$D$23:$D${22 + len(regions)}"
        elif getattr(category_source, "numRef", None) is not None:
            category_source.numRef.f = f"'Regional Performance'!$D$23:$D${22 + len(regions)}"
        chart.series[0].val.numRef.f = f"'Regional Performance'!$V$23:$V${22 + len(regions)}"


def _copy_sheet_with_charts(workbook, source, title: str):
    sheet = workbook.copy_worksheet(source)
    sheet.title = title
    sheet._charts = [copy(chart) for chart in source._charts]
    return sheet


def _set_chart_series(chart, sheet_name: str, category_row: int, value_row: int) -> None:
    if not chart.series:
        return
    series = chart.series[0]
    if getattr(series.cat, "numRef", None) is not None:
        series.cat.numRef.f = f"'{sheet_name}'!$E${category_row}:$P${category_row}"
    elif getattr(series.cat, "strRef", None) is not None:
        series.cat.strRef.f = f"'{sheet_name}'!$E${category_row}:$P${category_row}"
    series.val.numRef.f = f"'{sheet_name}'!$E${value_row}:$P${value_row}"


def populate_region_sheets(workbook, hierarchy: OrderedDict[str, Store], history: list[HistoryRecord],
                           current: list[dict[str, Any]], months: list[date],
                           regions: list[str], report_month: date) -> None:
    existing = {sheet.title: sheet for sheet in workbook.worksheets if re.fullmatch(r"Region .+", sheet.title)}
    template = existing.get("Region 1") or next(iter(existing.values()), None)
    if template is None:
        raise ReportGenerationError("The LIVE report contains no Region template tab.")
    desired_names = {f"Region {region}" for region in regions}
    for region in regions:
        name = f"Region {region}"
        sheet = workbook[name] if name in workbook.sheetnames else _copy_sheet_with_charts(workbook, template, name)
        _strip_formulas(sheet)
        stores = [store for store in hierarchy.values() if store.status == "Active" and store.region == region]
        legacy_totals_start = _find_row(sheet, "Total Count", 1, 7) or 7
        legacy_current_header = None
        for row in range(7, sheet.max_row + 1):
            if _norm(sheet.cell(row, 1).value) == "store code":
                legacy_current_header = row
                break
        legacy_current_totals = _find_row(
            sheet, "Total tests", 1, (legacy_current_header or 6) + 1
        ) or 1
        store_end = 6 + len(stores)
        totals_start = max(legacy_totals_start, store_end + 2)
        current_title_row = totals_start + 6
        current_header_row = current_title_row + 3
        current_start = current_header_row + 1
        current_end = current_start + len(stores) - 1
        current_totals = max(legacy_current_totals, current_end + 3)
        needed = current_totals + 2
        _ensure_rows(sheet, needed, 7, 16)
        for merged_range in list(sheet.merged_cells.ranges):
            sheet.unmerge_cells(str(merged_range))
        sheet.merge_cells("C3:I3")
        sheet.merge_cells(start_row=current_title_row, start_column=3,
                          end_row=current_title_row, end_column=9)
        for merged_row in range(current_header_row, current_end + 1):
            sheet.merge_cells(start_row=merged_row, start_column=4,
                              end_row=merged_row, end_column=5)
            sheet.merge_cells(start_row=merged_row, start_column=9,
                              end_row=merged_row, end_column=11)
        _clear_values(sheet, 3, max(sheet.max_row, needed), 1, 16)
        sheet["A3"] = name
        sheet["C3"] = f"External Test Purchases {report_month:%B %Y}"
        for offset, month in enumerate(months, 5):
            month_records = _records_in_month(history, month)
            orders = Counter(record.order for record in month_records if record.order)
            sheet.cell(5, offset).value = orders.most_common(1)[0][0] if orders else ""
            sheet.cell(6, offset).value = datetime(month.year, month.month, 10)
        for column, value in enumerate(["Store Code", "Store Name", "Police", "Total"], 1):
            sheet.cell(6, column).value = value
        for row, store in enumerate(stores, 7):
            sheet.cell(row, 1).value = _excel_code(store.code)
            sheet.cell(row, 2).value = store.name
            failures = 0
            for column, month in enumerate(months, 5):
                records = [record for record in _records_in_month(history, month)
                           if record.store_code == store.code and _result(record.result)]
                value = "-"
                if records:
                    failures += sum(_result(record.result) == "fail" for record in records)
                    value = "fail" if any(_result(record.result) == "fail" for record in records) else "P"
                sheet.cell(row, column).value = value
            sheet.cell(row, 4).value = failures
            sheet.cell(row, 4).number_format = "0"
        for index, label in enumerate(["Total Count", "Total Passes ", "Total Fails", "Pass Rate"]):
            row = totals_start + index
            sheet.cell(row, 1).value = label
            for column, month in enumerate(months, 5):
                records = [record for record in _records_in_month(history, month)
                           if record.store_code in {store.code for store in stores} and _result(record.result)]
                passes = sum(_result(record.result) == "pass" for record in records)
                fails = sum(_result(record.result) == "fail" for record in records)
                sheet.cell(row, column).value = [passes + fails, passes, fails,
                                                 passes / (passes + fails) * 100 if passes + fails else "-"][index]
                sheet.cell(row, column).number_format = "0"
        sheet.cell(current_title_row, 3).value = f"External Test Purchases {report_month:%B %Y}"
        sheet.cell(current_title_row + 1, 1).value = name
        for column, value in enumerate(["Store Code", "Store Name", "Date", "Approx time", "", "Result"], 1):
            cell = sheet.cell(current_header_row, column)
            if cell.__class__.__name__ != "MergedCell":
                cell.value = value
        current_by = defaultdict(list)
        for record in current:
            current_by[_normalise_code(record["store_code"])].append(record)
        for row, store in enumerate(stores, current_start):
            records = current_by[store.code]
            sheet.cell(row, 1).value = _excel_code(store.code)
            sheet.cell(row, 2).value = store.name
            if records:
                selected = sorted(records, key=lambda record: record["visit_date"] or datetime.min)[0]
                sheet.cell(row, 3).value = selected["visit_date"]
                sheet.cell(row, 4).value = selected["visit_time"]
                sheet.cell(row, 6).value = "fail" if any(_result(record["result"]) == "fail" for record in records) else "P"
            else:
                sheet.cell(row, 3).value = "-"
                sheet.cell(row, 4).value = "-"
                sheet.cell(row, 6).value = "-"
            sheet.cell(row, 3).number_format = "dd/mm/yy"
            sheet.cell(row, 4).number_format = "h:mm AM/PM"
        region_current = [record for record in current if _normalise_code(record["store_code"]) in {store.code for store in stores}]
        completed = [record for record in region_current if _result(record["result"])]
        passes = sum(_result(record["result"]) == "pass" for record in completed)
        for offset, (label, value) in enumerate([
            ("Total tests ", len(completed)), ("Total passes ", passes),
            ("", passes / len(completed) * 100 if completed else "-"),
        ]):
            row = current_totals + offset
            sheet.cell(row, 1).value = label
            sheet.cell(row, 2).value = "="
            sheet.cell(row, 2).data_type = "s"
            sheet.cell(row, 3).value = value
            sheet.cell(row, 3).number_format = "0"
            if offset == 2:
                sheet.cell(row, 4).value = "%"
        if sheet._charts:
            _set_chart_series(sheet._charts[0], name, 6, totals_start + 3)
        sheet.auto_filter.ref = f"A6:P{max(6, store_end)}"
    for name in list(existing):
        if name not in desired_names and name in workbook.sheetnames:
            workbook.remove(workbook[name])


def populate_public_reports(workbook, current: list[dict[str, Any]], history: list[HistoryRecord],
                            ytd: list[HistoryRecord], hierarchy: OrderedDict[str, Store],
                            regions: list[str], months: list[date], report_month: date) -> None:
    populate_summary(workbook["Summary Data"], current)
    populate_historic(workbook["Historic Data"], history)
    populate_store_performance(
        workbook["Store Performance"], workbook["Store Performance (2)"], hierarchy,
        current, ytd, report_month,
    )
    populate_self_scan(workbook["Self-Scan Performance"], hierarchy, history, months)
    populate_regional_performance(
        workbook["Regional Performance"], workbook["Regional P.Graphs"], hierarchy,
        history, months, regions,
    )
    populate_region_sheets(workbook, hierarchy, history, current, months, regions, report_month)

    existing_postcodes = [
        _text(workbook["Postcode Performance"].cell(row, 1).value)
        for row in range(7, workbook["Postcode Performance"].max_row + 1)
        if _text(workbook["Postcode Performance"].cell(row, 1).value) not in {"", "Total"}
    ]
    observed = {_postcode_region(record["postcode"]) for record in current}
    observed.update(_postcode_region(record.postcode) for record in ytd)
    postcode_labels = [label for label in existing_postcodes if label in observed]
    postcode_labels.extend(sorted(observed - set(postcode_labels)))
    populate_metric_sheet(
        workbook["Postcode Performance"], postcode_labels, current, ytd,
        lambda record: _postcode_region(record["postcode"]),
        lambda record: _postcode_region(record.postcode), report_month,
    )
    till_labels = ["Normal Till", "Self-Scan Till", "Kiosk"]
    populate_metric_sheet(
        workbook["Till Type Performance"], till_labels, current, ytd,
        lambda record: _normalise_till(record["till_type"]),
        lambda record: _normalise_till(record.till_type), report_month,
    )
    day_labels = ["Mon", "Tues", "Wed", "Thur", "Fri", "Sat", "Sun"]
    populate_metric_sheet(
        workbook["Day of Week Performance"], day_labels, current, ytd,
        lambda record: _day_label(record["visit_date"]),
        lambda record: _day_label(record.visit_date), report_month,
    )
    hour_labels = list(range(24))
    populate_metric_sheet(
        workbook["Time of Day Performance"], hour_labels, current, ytd,
        lambda record: _hour(record["visit_time"]),
        lambda record: _hour(record.visit_time), report_month,
    )
    populate_performance_over_time(workbook["Performance over Time"], history, months)


def _set_summary_formulas(workbook, current_last: int) -> None:
    sheet = workbook["Summary Data"]
    last = _last_data_row(sheet, 1, 8)
    end = max(8, last)
    sheet["B3"] = f"=COUNTA($A$8:$A${end})"
    sheet["B4"] = f'=COUNTIF($G$8:$G${end},"PASS")+COUNTIF($G$8:$G${end},"FAIL")'
    sheet["B5"] = '=IF($B$4=0,"-",COUNTIF($G$8:$G$%d,"PASS")/$B$4)' % end
    sheet["X3"] = "='Checks'!$B$25"
    formulas = [
        "='This Period'!W{row}", "=T('This Period'!H{row})",
        "=T('This Period'!M{row})", "='This Period'!Q{row}",
        "='This Period'!IL{row}", "='This Period'!R{row}",
        "=UPPER('This Period'!T{row})", "=T('This Period'!AG{row})",
        "=T('This Period'!AH{row})", "=T('This Period'!AI{row})",
        "=T('This Period'!AJ{row})", "=T('This Period'!AK{row})",
        "=T('This Period'!AL{row})", "=T('This Period'!AN{row})",
        "=T('This Period'!AO{row})", "=T('This Period'!AP{row})",
        "=T('This Period'!AQ{row})", "='This Period'!AR{row}",
        "=T('This Period'!AX{row})", "='This Period'!AZ{row}",
        "='This Period'!BE{row}", "=T('This Period'!BG{row})",
        "=T('This Period'!BD{row})", "=T('This Period'!BF{row})",
    ]
    for target_row, source_row in zip(range(8, last + 1), range(4, current_last + 1)):
        for column, formula in enumerate(formulas, 1):
            sheet.cell(target_row, column).value = formula.format(row=source_row)


def _set_historic_formulas(workbook, cumulative_last: int) -> None:
    sheet = workbook["Historic Data"]
    last = _last_data_row(sheet, 1, 4)
    for row, source_row in zip(range(4, last + 1), range(4, cumulative_last + 1)):
        sheet.cell(row, 1).value = f"='Cumulative'!V{source_row}"
        sheet.cell(row, 2).value = f'=IFERROR(VLOOKUP(A{row},\'StoreList\'!$A:$B,2,FALSE),"")'
        sheet.cell(row, 3).value = f"='Cumulative'!K{source_row}"
        sheet.cell(row, 4).value = f"='Cumulative'!P{source_row}"
        sheet.cell(row, 5).value = f"=UPPER('Cumulative'!S{source_row})"


def _set_store_performance_formulas(workbook, current_last: int, ytd_last: int) -> None:
    sheet = workbook["Store Performance"]
    total_row = _find_row(sheet, "Total", 1, 7)
    if total_row is None:
        return
    last_store = total_row - 2
    for row in range(7, last_store + 1):
        current_key = f"'This Period'!$IR$4:$IV${max(4, current_last)}"
        ytd_key = f"'YTD'!$IV$4:$IV${max(4, ytd_last)}"
        sheet.cell(row, 3).value = f'=COUNTIF({current_key},$A{row}&"PASS")+COUNTIF({current_key},$A{row}&"FAIL")+COUNTIF({current_key},$A{row}&"ABORT")'
        sheet.cell(row, 4).value = f'=COUNTIF({current_key},$A{row}&"PASS")+COUNTIF({current_key},$A{row}&"FAIL")'
        sheet.cell(row, 5).value = f'=COUNTIF({current_key},$A{row}&"FAIL")'
        sheet.cell(row, 6).value = f'=COUNTIF({current_key},$A{row}&"PASS")'
        sheet.cell(row, 7).value = f'=IF(D{row}=0,"-",F{row}/D{row})'
        sheet.cell(row, 8).value = f'=COUNTIF({ytd_key},$A{row}&"PASS")+COUNTIF({ytd_key},$A{row}&"FAIL")+COUNTIF({ytd_key},$A{row}&"ABORT")'
        sheet.cell(row, 9).value = f'=COUNTIF({ytd_key},$A{row}&"PASS")+COUNTIF({ytd_key},$A{row}&"FAIL")'
        sheet.cell(row, 10).value = f'=COUNTIF({ytd_key},$A{row}&"FAIL")'
        sheet.cell(row, 11).value = f'=COUNTIF({ytd_key},$A{row}&"PASS")'
        sheet.cell(row, 12).value = f'=IF(I{row}=0,"-",K{row}/I{row})'
    for column in [3, 4, 5, 6, 8, 9, 10, 11]:
        letter = get_column_letter(column)
        sheet.cell(total_row, column).value = f"=SUM({letter}7:{letter}{last_store})"
    sheet.cell(total_row, 7).value = f'=IF(D{total_row}=0,"-",F{total_row}/D{total_row})'
    sheet.cell(total_row, 12).value = f'=IF(I{total_row}=0,"-",K{total_row}/I{total_row})'

    league = workbook["Store Performance (2)"]
    last = _last_data_row(league, 1, 6)
    key = f"'YTD'!$IV$4:$IV${max(4, ytd_last)}"
    for row in range(6, last + 1):
        league.cell(row, 3).value = f'=COUNTIF({key},$A{row}&"FAIL")'


def _set_self_scan_formulas(workbook, cumulative_last: int) -> None:
    sheet = workbook["Self-Scan Performance"]
    right_table = next((table for table in sheet.tables.values() if table.name == "Table1"), None)
    totals_table = next((table for table in sheet.tables.values() if table.name == "Table2"), None)
    if right_table is None or totals_table is None:
        return
    _, _, _, right_total = range_boundaries(right_table.ref)
    last_store = right_total - 1
    _, total_start, _, total_end = range_boundaries(totals_table.ref)
    for offset, column in enumerate(range(3, 15), 18):
        letter = get_column_letter(column)
        sheet.cell(4, column).value = f"='Input'!$A${offset}"
        sheet.cell(5, column).value = f"='Input'!$B${offset}"
        for row in range(6, last_store + 1):
            sheet.cell(row, column).value = (
                f'=IFERROR(VLOOKUP($A{row}&{letter}$4,\'Cumulative\'!$IB$4:$IC${max(4, cumulative_last)},2,FALSE),"-")'
            )
    for row in range(6, last_store + 1):
        sheet.cell(row, 18).value = f'=IF(COUNTIF(C{row}:N{row},"P")+COUNTIF(C{row}:N{row},"Fail")=0,"-",COUNTIF(C{row}:N{row},"P")+COUNTIF(C{row}:N{row},"Fail"))'
        sheet.cell(row, 19).value = f'=IF(R{row}="-","-",COUNTIF(C{row}:N{row},"P"))'
        sheet.cell(row, 20).value = f'=IF(R{row}="-","-",S{row}/R{row})'
    sheet.cell(right_total, 18).value = f'=IF(SUM(R6:R{last_store})=0,"-",SUM(R6:R{last_store}))'
    sheet.cell(right_total, 19).value = f'=IF(R{right_total}="-","-",SUM(S6:S{last_store}))'
    sheet.cell(right_total, 20).value = f'=IF(R{right_total}="-","-",S{right_total}/R{right_total})'
    for column in range(3, 15):
        letter = get_column_letter(column)
        sheet.cell(total_start, column).value = f'=IF(COUNTIF({letter}6:{letter}{last_store},"P")+COUNTIF({letter}6:{letter}{last_store},"Fail")=0,"-",COUNTIF({letter}6:{letter}{last_store},"P")+COUNTIF({letter}6:{letter}{last_store},"Fail"))'
        sheet.cell(total_start + 1, column).value = f'=IF({letter}{total_start}="-","-",COUNTIF({letter}6:{letter}{last_store},"P"))'
        sheet.cell(total_start + 2, column).value = f'=IF({letter}{total_start}="-","-",{letter}{total_start + 1}/{letter}{total_start})'


def _set_metric_formulas(workbook, current_last: int, ytd_last: int) -> None:
    helpers = {
        "Postcode Performance": ("IQ", "IU"),
        "Till Type Performance": ("IE", "IE"),
        "Day of Week Performance": ("IO", "IS"),
        "Time of Day Performance": ("IJ", "IL"),
    }
    for name, (current_helper, ytd_helper) in helpers.items():
        sheet = workbook[name]
        total_row = _find_row(sheet, "Total", 1, 7)
        if total_row is None:
            continue
        last_label = total_row - 2
        current_key = f"'This Period'!${current_helper}$4:${current_helper}${max(4, current_last)}"
        ytd_key = f"'YTD'!${ytd_helper}$4:${ytd_helper}${max(4, ytd_last)}"
        for row in range(7, last_label + 1):
            sheet.cell(row, 2).value = f'=COUNTIF({current_key},$A{row}&"PASS")+COUNTIF({current_key},$A{row}&"FAIL")+COUNTIF({current_key},$A{row}&"ABORT")'
            sheet.cell(row, 3).value = f'=COUNTIF({current_key},$A{row}&"PASS")+COUNTIF({current_key},$A{row}&"FAIL")'
            sheet.cell(row, 4).value = f'=COUNTIF({current_key},$A{row}&"FAIL")'
            sheet.cell(row, 5).value = f'=COUNTIF({current_key},$A{row}&"PASS")'
            sheet.cell(row, 6).value = f'=IF(C{row}=0,"-",E{row}/C{row})'
            sheet.cell(row, 7).value = f'=COUNTIF({ytd_key},$A{row}&"PASS")+COUNTIF({ytd_key},$A{row}&"FAIL")+COUNTIF({ytd_key},$A{row}&"ABORT")'
            sheet.cell(row, 8).value = f'=COUNTIF({ytd_key},$A{row}&"PASS")+COUNTIF({ytd_key},$A{row}&"FAIL")'
            sheet.cell(row, 9).value = f'=COUNTIF({ytd_key},$A{row}&"FAIL")'
            sheet.cell(row, 10).value = f'=COUNTIF({ytd_key},$A{row}&"PASS")'
            sheet.cell(row, 11).value = f'=IF(H{row}=0,"-",J{row}/H{row})'
        for column in [2, 3, 4, 5, 7, 8, 9, 10]:
            letter = get_column_letter(column)
            sheet.cell(total_row, column).value = f"=SUM({letter}7:{letter}{last_label})"
        sheet.cell(total_row, 6).value = f'=IF(C{total_row}=0,"-",E{total_row}/C{total_row})'
        sheet.cell(total_row, 11).value = f'=IF(H{total_row}=0,"-",J{total_row}/H{total_row})'


def _set_regional_formulas(workbook, cumulative_last: int, regions: list[str]) -> None:
    sheet = workbook["Regional Performance"]
    region_rows: dict[int, list[int]] = {}
    for block, data_start in [(0, 7), (1, 23)]:
        header_row = 4 + block * 16
        rows = list(range(data_start, data_start + len(regions)))
        region_rows[block] = rows
        total_row = data_start + len(regions)
        for row in rows:
            for column in range(5, 23, 3):
                pass_letter = get_column_letter(column)
                fail_letter = get_column_letter(column + 1)
                sheet.cell(row, column).value = f'=COUNTIF(\'Cumulative\'!$IH$4:$IH${max(4, cumulative_last)},$D{row}&{pass_letter}${header_row}&{pass_letter}$6)'
                sheet.cell(row, column + 1).value = f'=COUNTIF(\'Cumulative\'!$IH$4:$IH${max(4, cumulative_last)},$D{row}&{pass_letter}${header_row}&{fail_letter}$6)'
                sheet.cell(row, column + 2).value = f'=IF(SUM({pass_letter}{row}:{fail_letter}{row})=0,"",{pass_letter}{row}/SUM({pass_letter}{row}:{fail_letter}{row})*100)'
        for column in range(5, 23, 3):
            pass_letter = get_column_letter(column)
            fail_letter = get_column_letter(column + 1)
            rate_letter = get_column_letter(column + 2)
            sheet.cell(total_row, column).value = f"=SUM({pass_letter}{data_start}:{pass_letter}{total_row - 1})"
            sheet.cell(total_row, column + 1).value = f"=SUM({fail_letter}{data_start}:{fail_letter}{total_row - 1})"
            sheet.cell(total_row, column + 2).value = f'=IF(SUM({pass_letter}{total_row}:{fail_letter}{total_row})=0,"",{pass_letter}{total_row}/SUM({pass_letter}{total_row}:{fail_letter}{total_row})*100)'
    for index, region in enumerate(regions, 4):
        first_row = 7 + regions.index(region)
        second_row = 23 + regions.index(region)
        pass_cells = [f"{get_column_letter(column)}{row}" for row in [first_row, second_row] for column in range(5, 23, 3)]
        fail_cells = [f"{get_column_letter(column)}{row}" for row in [first_row, second_row] for column in range(6, 23, 3)]
        pass_sum = "+".join(pass_cells)
        fail_sum = "+".join(fail_cells)
        sheet.cell(40, index).value = f'=IF(({pass_sum}+{fail_sum})=0,"",({pass_sum})/({pass_sum}+{fail_sum}))'


def _set_region_formulas(workbook, current_last: int, cumulative_last: int,
                         regions: list[str]) -> None:
    for region in regions:
        sheet = workbook[f"Region {region}"]
        totals_start = _find_row(sheet, "Total Count", 1, 7)
        current_header = None
        for row in range((totals_start or 7) + 4, sheet.max_row + 1):
            if _norm(sheet.cell(row, 1).value) == "store code":
                current_header = row
                break
        current_totals = _find_row(sheet, "Total tests", 1, (current_header or 6) + 1)
        if totals_start is None or current_header is None or current_totals is None:
            continue
        sheet["C3"] = '="External Test Purchases "&\'Checks\'!$B$26'
        current_title_row = current_header - 3
        sheet.cell(current_title_row, 3).value = '="External Test Purchases "&\'Checks\'!$B$26'
        for offset, column in enumerate(range(5, 17), 18):
            letter = get_column_letter(column)
            sheet.cell(5, column).value = f"='Input'!$A${offset}"
            sheet.cell(6, column).value = f"='Input'!$B${offset}"
        upper_rows = [row for row in range(7, totals_start) if sheet.cell(row, 1).value not in (None, "")]
        for row in upper_rows:
            sheet.cell(row, 4).value = f'=COUNTIF(E{row}:P{row},"Fail")'
            for column in range(5, 17):
                letter = get_column_letter(column)
                sheet.cell(row, column).value = f'=IFERROR(VLOOKUP($A{row}&{letter}$5,\'Cumulative\'!$IF$4:$IL${max(4, cumulative_last)},2,FALSE),"-")'
        last_upper = upper_rows[-1] if upper_rows else 6
        for index in range(4):
            row = totals_start + index
            for column in range(5, 17):
                letter = get_column_letter(column)
                if index == 0:
                    formula = f'=COUNTIF({letter}7:{letter}{last_upper},"P")+COUNTIF({letter}7:{letter}{last_upper},"Fail")'
                elif index == 1:
                    formula = f'=COUNTIF({letter}7:{letter}{last_upper},"P")'
                elif index == 2:
                    formula = f'=COUNTIF({letter}7:{letter}{last_upper},"Fail")'
                else:
                    formula = f'=IF({letter}{totals_start}=0,"",{letter}{totals_start + 1}/{letter}{totals_start}*100)'
                sheet.cell(row, column).value = formula
        lower_rows = [
            row for row in range(current_header + 1, current_totals)
            if sheet.cell(row, 1).value not in (None, "")
        ]
        for row in lower_rows:
            sheet.cell(row, 3).value = f'=IFERROR(VLOOKUP($A{row},\'This Period\'!$A$4:$Q${max(4, current_last)},17,FALSE),"-")'
            sheet.cell(row, 4).value = f'=IFERROR(VLOOKUP($A{row},\'This Period\'!$A$4:$R${max(4, current_last)},18,FALSE),"-")'
            sheet.cell(row, 6).value = f'=IFERROR(VLOOKUP($A{row},\'This Period\'!$A$4:$IG${max(4, current_last)},241,FALSE),"-")'
        last_lower = lower_rows[-1] if lower_rows else current_header
        sheet.cell(current_totals, 3).value = f'=COUNTIF(F{current_header + 1}:F{last_lower},"P")+COUNTIF(F{current_header + 1}:F{last_lower},"Fail")'
        sheet.cell(current_totals + 1, 3).value = f'=COUNTIF(F{current_header + 1}:F{last_lower},"P")'
        sheet.cell(current_totals + 2, 3).value = f'=IF(C{current_totals}=0,"-",C{current_totals + 1}/C{current_totals}*100)'


def _set_performance_over_time_formulas(workbook, regions: list[str]) -> None:
    sheet = workbook["Performance over Time"]
    totals = [7 + len(regions), 23 + len(regions)]
    for index, column in enumerate(range(2, 14)):
        input_row = 18 + index
        block = 0 if index < 6 else 1
        metric_column = 5 + (index % 6) * 3
        pass_letter = get_column_letter(metric_column)
        fail_letter = get_column_letter(metric_column + 1)
        total_row = totals[block]
        sheet.cell(5, column).value = f'=TEXT(\'Input\'!$B${input_row},"mmm yyyy")'
        sheet.cell(6, column).value = f'=IF((\'Regional Performance\'!{pass_letter}{total_row}+\'Regional Performance\'!{fail_letter}{total_row})=0,"-",\'Regional Performance\'!{pass_letter}{total_row}/(\'Regional Performance\'!{pass_letter}{total_row}+\'Regional Performance\'!{fail_letter}{total_row}))'


def apply_public_formulas(workbook, current_last: int, cumulative_last: int,
                          ytd_last: int, regions: list[str]) -> None:
    """Restore the formula-driven public tabs used by the internal LIVE file."""
    _set_summary_formulas(workbook, current_last)
    _set_historic_formulas(workbook, cumulative_last)
    _set_store_performance_formulas(workbook, current_last, ytd_last)
    _set_self_scan_formulas(workbook, cumulative_last)
    _set_metric_formulas(workbook, current_last, ytd_last)
    _set_regional_formulas(workbook, cumulative_last, regions)
    _set_region_formulas(workbook, current_last, cumulative_last, regions)
    _set_performance_over_time_formulas(workbook, regions)


def _remove_invalid_names(workbook, removed_sheets: set[str] | None = None) -> None:
    removed_sheets = removed_sheets or set()
    containers = [workbook.defined_names]
    containers.extend(sheet.defined_names for sheet in workbook.worksheets)
    for container in containers:
        for name, defined_name in list(container.items()):
            reference = _text(getattr(defined_name, "attr_text", ""))
            if "#REF!" in reference or re.search(r"\[\d+\]", reference) or any(
                f"'{sheet}'!" in reference or f"{sheet}!" in reference for sheet in removed_sheets
            ):
                del container[name]


def _strip_table_formulas(workbook) -> None:
    """Keep the non-LIVE report genuinely values-only, including table metadata."""
    for sheet in workbook.worksheets:
        for table in sheet.tables.values():
            for column in table.tableColumns:
                column.calculatedColumnFormula = None
                column.totalsRowFormula = None
                column.totalsRowFunction = None


def _remove_invalid_formulas(workbook) -> None:
    """Remove unusable legacy checks that already point at deleted template ranges."""
    for sheet in workbook.worksheets:
        for cell in list(sheet._cells.values()):
            if cell.data_type == "f" and "#REF!" in str(cell.value):
                cell.value = None


def _select_sheet(workbook, title: str) -> None:
    for sheet in workbook.worksheets:
        sheet.sheet_view.tabSelected = False
    target = workbook[title]
    workbook.active = workbook._sheets.index(target)
    target.sheet_view.tabSelected = True


def _save(workbook) -> bytes:
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def generate_reports(csv_bytes: bytes, previous_live_bytes: bytes,
                     store_bytes: bytes) -> GenerationResult:
    current, report_month, product_word, stats, warnings = map_audit_export(csv_bytes)
    workbook = _load_live_workbook(previous_live_bytes)
    database = read_store_database(store_bytes)
    previous = extract_previous_stores(workbook)
    hierarchy, hierarchy_warnings = build_hierarchy(database, previous, current)
    warnings.extend(hierarchy_warnings)

    update_store_list(workbook["StoreList"], hierarchy)
    current_last = update_this_period(workbook["This Period"], current)
    cumulative_last = update_cumulative(workbook["Cumulative"], current)
    history, ytd = extract_history(workbook["Cumulative"], report_month.year)
    ytd_last = update_ytd(workbook["YTD"], ytd)
    months = update_input(workbook["Input"], history, current, report_month)
    workbook["Checks"]["B26"] = f"{report_month:%B %Y}"
    regions = sorted({store.region for store in database.values()}, key=_natural_key)
    if any(store.status == "Unmapped" for store in hierarchy.values()):
        regions.append("Unmapped")
    populate_public_reports(workbook, current, history, ytd, hierarchy, regions, months, report_month)

    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.calculation.calcMode = "auto"
    workbook._external_links = []
    _remove_invalid_formulas(workbook)
    _remove_invalid_names(workbook)
    static_bytes = _save(workbook)

    client = load_workbook(io.BytesIO(static_bytes), data_only=False, keep_links=False)
    public_sheets = set(PUBLIC_BASE_SHEETS) | {f"Region {region}" for region in regions}
    removed = {name for name in client.sheetnames if name not in public_sheets}
    for name in list(client.sheetnames):
        if name not in public_sheets:
            client.remove(client[name])
    for sheet in client.worksheets:
        _strip_formulas(sheet)
    _strip_table_formulas(client)
    client._external_links = []
    _remove_invalid_names(client, removed)
    _select_sheet(client, "Summary Data")
    client_bytes = _save(client)
    client.close()

    apply_public_formulas(workbook, current_last, cumulative_last, ytd_last, regions)
    _remove_invalid_formulas(workbook)
    _remove_invalid_names(workbook)
    _select_sheet(workbook, "This Period")
    live_bytes = _save(workbook)
    workbook.close()

    completed = [record for record in current if _result(record["result"])]
    current_passes = sum(_result(record["result"]) == "pass" for record in completed)
    ytd_completed = [record for record in ytd if _result(record.result)]
    ytd_passes = sum(_result(record.result) == "pass" for record in ytd_completed)
    earlier_months = sorted({
        _month_start(record.visit_date) for record in history
        if record.visit_date and record.visit_date.date() < report_month and _result(record.result)
    })
    previous_month = earlier_months[-1] if earlier_months else None
    previous_records = _records_in_month(history, previous_month) if previous_month else []
    previous_metric = _metric([record for record in previous_records if _result(record.result)])
    stats.update({
        "completed_visits": len(completed),
        "pass_rate": current_passes / len(completed) if completed else None,
        "ytd_completed_visits": len(ytd_completed),
        "ytd_pass_rate": ytd_passes / len(ytd_completed) if ytd_completed else None,
        "previous_reporting_month": previous_month,
        "previous_pass_rate": previous_metric[4] if previous_metric[4] != "-" else None,
        "database_stores": len(database),
        "report_stores": len(hierarchy),
        "closed_history_stores": sum(store.status == "Closed" for store in hierarchy.values()),
        "regions": len(regions),
        "current_last_row": current_last,
        "cumulative_last_row": cumulative_last,
        "ytd_last_row": ytd_last,
    })

    live_name = f"Scotmid Test Purchases {report_month:%B %Y} - LIVE.xlsx"
    client_name = f"Scotmid Test Purchases {report_month:%B %Y}.xlsx"
    archive_buffer = io.BytesIO()
    with zipfile.ZipFile(archive_buffer, "w", zipfile.ZIP_DEFLATED) as archive:
        archive.writestr(live_name, live_bytes)
        archive.writestr(client_name, client_bytes)
    return GenerationResult(
        live_bytes, client_bytes, archive_buffer.getvalue(), live_name, client_name,
        report_month, product_word, stats, warnings,
    )


def _uk_greeting(now: datetime | None = None) -> str:
    now = now or datetime.now(ZoneInfo("Europe/London"))
    if 5 <= now.hour < 12:
        return "Morning"
    if 12 <= now.hour < 17:
        return "Afternoon"
    return "Evening"


def _format_rate(value: float | None, decimals: int = 0) -> str:
    return "N/A" if value is None else f"{value * 100:.{decimals}f}%"


def build_email_text(generated: GenerationResult, benchmark_rate: float | None,
                     now: datetime | None = None) -> str:
    pass_rate = generated.stats.get("pass_rate")
    previous_rate = generated.stats.get("previous_pass_rate")
    previous_month = generated.stats.get("previous_reporting_month")
    difference = None if pass_rate is None or previous_rate is None else (pass_rate - previous_rate) * 100
    if difference is None:
        comparison = "A comparison with the previous reporting month is not available"
    elif abs(difference) < 0.5:
        comparison = "This is broadly unchanged from last month"
    else:
        direction = "increase" if difference > 0 else "decrease"
        article = "an" if direction == "increase" else "a"
        period = "last month" if previous_month == _add_months(generated.report_month, -1) else "the previous reporting month"
        comparison = f"This is {article} {direction} of c. {abs(difference):.0f}% from {period}"
    benchmark_text = (
        f"the off-trade {generated.product_word} pass rate for {generated.report_month:%B} was {benchmark_rate:.0f}%."
        if benchmark_rate is not None
        else f"the off-trade {generated.product_word} pass rate for {generated.report_month:%B} was [enter benchmark]%."
    )
    return (
        f"{_uk_greeting(now)} Ian & Kevin,\n\n"
        "Please find attached the Serve Legal report detailing visits completed in "
        f"{generated.report_month:%B}.\n\n"
        f"As you’ll see from the report, your pass rate was {_format_rate(pass_rate)} based on "
        f"{generated.stats['completed_visits']} completed AV audits. {comparison} and puts your "
        f"year-to-date pass rate at {_format_rate(generated.stats.get('ytd_pass_rate'), 1)}.\n\n"
        f"To give you some context within the market, {benchmark_text}"
    )


def run_app() -> None:
    import streamlit as st

    st.set_page_config(page_title="Scotmid Report Generator", page_icon="📊", layout="wide")
    st.title("Scotmid Report Generator")
    st.caption(f"Generator version {GENERATOR_VERSION}")
    st.write(
        "Upload the new audit export, the previous LIVE report and the current Store DB. "
        "The generator produces both the updated LIVE workbook and the client-facing workbook."
    )
    with st.expander("What the generator updates"):
        st.markdown(
            "- Maps the new Alcohol, Cigarettes or E-Cig audit export and removes aborts.\n"
            "- Replaces **This Period**, appends/de-duplicates **Cumulative**, and rebuilds **YTD**.\n"
            "- Refreshes **StoreList**, store names, region assignments, new sites and historical closures.\n"
            "- Rebuilds **Regional Performance**, **Regional P.Graphs**, and every **Region x** tab from the Store DB.\n"
            "- Produces an internal LIVE workbook and a values-only client workbook."
        )

    col1, col2, col3 = st.columns(3)
    with col1:
        audit_upload = st.file_uploader("1. New audit data export", type=["csv"])
    with col2:
        live_upload = st.file_uploader("2. Previous LIVE report", type=["xlsx"])
    with col3:
        store_upload = st.file_uploader("3. Current Store DB", type=["xlsx", "xlsm"])

    benchmark_rate = st.number_input(
        "Off-trade benchmark pass rate for the reporting month (%)",
        min_value=0.0, max_value=100.0, value=None, step=0.1,
        help="This market benchmark is used only in the generated email text.",
    )
    ready = all(upload is not None for upload in (audit_upload, live_upload, store_upload))
    payloads = None
    signature = None
    if ready:
        payloads = (audit_upload.getvalue(), live_upload.getvalue(), store_upload.getvalue())
        signature = tuple(
            (upload.name, len(payload), hashlib.sha256(payload).hexdigest())
            for upload, payload in zip((audit_upload, live_upload, store_upload), payloads)
        )
    if st.session_state.get("scotmid_input_signature") != signature:
        st.session_state.pop("scotmid_generated_report", None)
        st.session_state["scotmid_input_signature"] = signature

    if st.button("Generate reports", type="primary", disabled=not ready, use_container_width=True):
        try:
            with st.spinner("Generating and validating the Scotmid reports..."):
                generated = generate_reports(*payloads)
            st.session_state["scotmid_generated_report"] = generated
        except ReportGenerationError as exc:
            st.session_state.pop("scotmid_generated_report", None)
            st.error(str(exc))
        except Exception as exc:
            st.session_state.pop("scotmid_generated_report", None)
            st.exception(exc)

    generated = st.session_state.get("scotmid_generated_report")
    if generated is not None:
        st.success(f"{generated.report_month:%B %Y} reports generated successfully.")
        columns = st.columns(5)
        for column, (label, key) in zip(columns, [
            ("Included visits", "included_rows"), ("Pass rate", "pass_rate"),
            ("YTD pass rate", "ytd_pass_rate"), ("Store DB sites", "database_stores"),
            ("Regions", "regions"),
        ]):
            value = generated.stats[key]
            column.metric(label, _format_rate(value, 1) if "rate" in key else value)
        st.caption(
            f"Removed {generated.stats['aborts_removed']} abort row(s) and "
            f"{generated.stats['unrelated_rows_removed']} unrelated row(s). "
            f"Retained {generated.stats['closed_history_stores']} historical closed store(s)."
        )
        for warning in generated.warnings:
            st.warning(warning)

        st.download_button(
            "Download both reports (.zip)", generated.zip_bytes,
            file_name=f"Scotmid Reports {generated.report_month:%B %Y}.zip",
            mime="application/zip", type="primary", use_container_width=True,
        )
        left, right = st.columns(2)
        left.download_button(
            "Download LIVE report", generated.live_bytes, generated.live_name,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
        right.download_button(
            "Download client report", generated.client_bytes, generated.client_name,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
        st.markdown("###### Email Text")
        st.code(build_email_text(generated, benchmark_rate), language="text")


if __name__ == "__main__":
    run_app()
